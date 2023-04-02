from tkinter import *
from tkinter import ttk
from tkinter.messagebox import *
import table2 as table2
from sqlite3 import *
from openpyxl import Workbook, load_workbook

# Получение информации из первой таблицы в бд
def information():
    with connect('database\database.db') as db:
        cursor = db.cursor()
        cursor.execute("SELECT * FROM table1")
        return cursor.fetchall()


# функция добавления новых записей
def form_submit():
    name = f_name.get() #Entry имя
    insert_inf = (name,)
    with connect('database\database.db') as db:
        cursor = db.cursor()
        query = """ INSERT INTO table1(group_name) VALUES (?)"""
        cursor.execute(query, insert_inf)
        db.commit()
        refresh()


# функция обновления таблицы1
def refresh():
    with connect('database\database.db') as db:
        cursor = db.cursor()
        cursor.execute(''' SELECT * FROM table1 ''')
        [table.delete(i) for i in
         table.get_children()]
        [table.insert('', 'end', values=row) for row in cursor.fetchall()]


# Функция удаления из таблицы1
def delete_user():
    with connect('database\database.db') as db:
        cursor = db.cursor()
        id = id_sel
        cursor.execute('''DELETE FROM table1 WHERE id = ?''', (id,))
        db.commit()
        refresh()


# Функция on_select
# при нажатии по ячейке Treview, создается глобальная переменная с id и column
def on_select(event):
    global id_sel
    global set_col
    id_sel = table.item(table.focus())
    id_sel = id_sel.get('values')[0]
    col = table.identify_column(event.x)
    set_col = table.column(col)
    set_col = set_col.get('id')
    if set_col == 'Группа':
        set_col = 'group_name'


# Функция изменения таблицы1
def changeDB():
    with connect('database\database.db') as db:
        cursor = db.cursor()
        id = id_sel
        whatchange = f_change.get()
        if set_col != 'id':
            cursor.execute("""Update table1 set""" + ' ' + set_col + """ = ? where id = ? """, (whatchange, id))
            db.commit()
            refresh()


# Функция показа справки
def show_info():
    msg = 'Кнопка "Добавить" работает следующим образом: указать имя, платеж, после нажатия создает и добавляет в файл\n\n' \
          'Кнопка "Изменить" работает следующим образом: Сначала надо нажать на изменяемую ячейку, где изменять столбец name, после указать новый name\n\n' \
          'Кнопка "Удалить" работает следующим образом: Сначала надо нажать на изменяемую ячейку, где удалить строку, после нажать кнопку\n\n'\
          'Кнопка "Вхождение в контекстном меню" работает следущющим образом: Сначала надо выбрать какое вхождение будем проверять,' \
          'после по каким столбцам из обеих таблиц, при нажатии кнопки проверка, на Label выдаст результат вхождения в другую таблицу\n\n'\
          'Кнопка "Выгрузка из бд" работает следующим образом: при нажатии "выгрузка в Excel" выгружает всю таблицу в папку database в формате xlsx'

    showinfo("Информация", msg)


# Функция вывода всей бд в excel в папку database
def db_to_excel():
    # Установка соединения с базой данных SQLite
    conn = connect('database\database.db')
    cur = conn.cursor()

    # Получение данных из базы данных в виде списка
    cur.execute("SELECT * FROM table1")
    results = cur.fetchall()

    # Создание объекта Workbook и листа для записи данных
    wb = Workbook()
    del wb['Sheet']

    # создание листов с именами 'table1' и 'table2'
    wb.create_sheet(title='table1')
    wb.create_sheet(title='table2')
    ws = wb.worksheets[0]
    sheet_names = wb.sheetnames
    # Запись заголовков столбцов в первую строку таблицы
    column_names = [description[0] for description in cur.description]
    ws.append(column_names)

    # Запись данных в таблицу Excel
    for row in results:
        ws.append(row)

    # Запись заголовков столбцов в первую строку листа2
    book2 = wb.worksheets[sheet_names.index('table2')]
    wb.active = book2
    ws = wb.active

    cur.execute("SELECT * FROM table2")
    results = cur.fetchall()
    column_names = [description[0] for description in cur.description]
    ws.append(column_names)

    # Запись данных в таблицу Excel
    for row in results:
        ws.append(row)

    # Сохранение таблицы в файл Excel
    wb.save("database\database.xlsx")

    # Закрытие соединения с базой данных SQLite
    cur.close()
    conn.close()


#  Главное окно
window = Tk()
window.title('subd')
window.minsize(700, 450)

frame_change = Frame(window, width=150, height=150, bg='white')  # блок для функционала субд
frame_view = Frame(window, width=150, height=150, bg='white')  # блок для просмотра базы данных
frame_change.place(relx=0, rely=0, relwidth=1, relheight=1)
frame_view.place(relx=0, rely=0.5, relwidth=1, relheight=0.5)

# порядок элементов
heads = ['id', 'Группа']
table = ttk.Treeview(frame_view, show='headings')  # дерево выполняющее свойство таблицы
table['columns'] = heads  # длина таблицы
table.bind('<ButtonRelease-1>', on_select)

# заголовки столбцов и их расположение
for header in heads:
    table.heading(header, text=header, anchor='center')
    table.column(header, anchor='center')

# добавление из бд в таблицу приложения
for row in information():
    table.insert('', END, values=row)
table.pack(expand=YES, fill=BOTH, side=LEFT)

# контекстное меню в Главном окне
mainmenu = Menu(window)
window.config(menu=mainmenu)


# Добавление кнопки "Выгрузка из бд" в контекстное меню
frombd_menu = Menu(mainmenu, tearoff=0)
mainmenu.add_cascade(label="Выгрузка из бд",
                     menu=frombd_menu)
frombd_menu.add_command(label='В Excel', command=db_to_excel)


# добавления новых имен в бд
l_name = ttk.Label(frame_change, text="Группа")
f_name = ttk.Entry(frame_change)
l_name.grid(row=0, column=0, sticky='w', padx=10, pady=10)
f_name.grid(row=0, column=1, sticky='w', padx=10, pady=10)

#  изменения бд
l_change = ttk.Label(frame_change, text="Заменить на:")
f_change = ttk.Entry(frame_change)  # entry на что меняем прошлое имя в бд
l_change.grid(row=1, column=0, sticky='w', padx=10, pady=10)
f_change.grid(row=1, column=1, sticky='w', padx=10, pady=10)

#  кнопка добавить
btn_submit = ttk.Button(frame_change, text="Добавить", command=form_submit)
btn_submit.grid(row=0, column=3, columnspan=2, sticky='w', padx=10, pady=10)

#  кнопка удалить
btn_delete = ttk.Button(frame_change, text="Удалить", command=delete_user)
btn_delete.grid(row=1, column=3, columnspan=2, sticky='w', padx=10, pady=10)

#  кнопка изменить
but_change = ttk.Button(frame_change, text='Изменить', command=changeDB)
but_change.grid(row=3, column=3, columnspan=2, sticky='w', padx=10, pady=10)

#  кнопка вызывающая справку
btn_reference = ttk.Button(frame_change, text="Справка", command=show_info)
btn_reference.grid(row=3, column=0, sticky='w', padx=10, pady=10)

# Кнопка вызова таблицы 2
create_new_table = ttk.Button(frame_change, text='Таблица 2', command=table2.create_table2)
create_new_table.grid(row=3, column=1, columnspan=2, sticky='w', padx=10, pady=10)

# скроллбар для treview
scrollpanel = ttk.Scrollbar(frame_view, command=table.yview)
table.configure(yscrollcommand=scrollpanel.set)
scrollpanel.pack(side=RIGHT, fill=Y)
table.pack(expand=YES, fill=BOTH)

window.mainloop()
